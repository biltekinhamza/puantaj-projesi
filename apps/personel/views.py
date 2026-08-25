from datetime import date, datetime
import base64
import mimetypes
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.db.models import Q
from django.http import HttpResponse
from io import BytesIO
from urllib.parse import quote
from django.shortcuts import get_object_or_404, redirect, render
from django.templatetags.static import static
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from apps.ayarlar.models import SirketBilgisi
from .forms import PersonelForm
from .models import Personel
from .selectors import tarihte_calisanlar
from .sozlesme import sgk_isyeri_sicil_no_from_departman


def _style_workbook(ws):
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 3, 45)
    ws.freeze_panes = "A2"


def _excel_response(wb, filename):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    safe_filename = filename.replace('"', "")
    response["Content-Disposition"] = (
        f'attachment; filename="{safe_filename}"; filename*=UTF-8\'\'{quote(safe_filename)}'
    )
    response["Content-Length"] = str(output.getbuffer().nbytes)
    response["Cache-Control"] = "no-store"
    return response

@login_required
def personel_ekle(request):
    form = PersonelForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("personel_liste")
    return render(request, "personel/personel_form.html", {"form": form})

@login_required
def personel_guncelle(request, pk):
    personel = get_object_or_404(Personel, pk=pk)
    form = PersonelForm(request.POST or None, instance=personel)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("personel_liste")
    return render(request, "personel/personel_form.html", {"form": form, "is_update": True, "personel": personel})

@login_required
def personel_liste(request):
    bugun = date.today()
    durum = request.GET.get("durum", "calisan")
    personel_turu = request.GET.get("personel_turu")
    departman = request.GET.get("departman")
    qs = Personel.objects.all().order_by("ad_soyad")

    if durum == "calisan":
        qs = qs.filter(ise_giris_tarihi__lte=bugun).filter(Q(isten_cikis_tarihi__isnull=True) | Q(isten_cikis_tarihi__gte=bugun))
    elif durum == "cikan":
        qs = qs.filter(isten_cikis_tarihi__lt=bugun)
    elif durum == "planli_cikis":
        qs = qs.filter(isten_cikis_tarihi__gte=bugun)

    if personel_turu:
        qs = qs.filter(personel_turu=personel_turu)
    if departman:
        qs = qs.filter(departman=departman)

    departmanlar = Personel.objects.exclude(departman__isnull=True).exclude(departman__exact="").values_list("departman", flat=True).distinct().order_by("departman")
    return render(request, "personel/personel_liste.html", {
        "personeller": qs,
        "durum": durum,
        "secili_personel_turu": personel_turu,
        "secili_departman": departman,
        "departmanlar": departmanlar,
        "toplam_sayisi": qs.count(),
        "bugun": bugun,
    })

@login_required
def personel_bilgi_raporu(request):
    return personel_liste(request)

@login_required
def personel_bilgi_excel(request):
    bugun = date.today()
    durum = request.GET.get("durum", "tum")
    qs = Personel.objects.all().order_by("ad_soyad")
    if durum == "calisan":
        qs = qs.filter(ise_giris_tarihi__lte=bugun).filter(Q(isten_cikis_tarihi__isnull=True) | Q(isten_cikis_tarihi__gte=bugun))
    elif durum == "cikan":
        qs = qs.filter(isten_cikis_tarihi__lt=bugun)
    elif durum == "planli_cikis":
        qs = qs.filter(isten_cikis_tarihi__gte=bugun)

    wb = Workbook()
    ws = wb.active
    ws.title = "Personel Bilgi"
    ws.append(["Ad Soyad", "Durum", "Personel Türü", "TC Kimlik", "Departman", "Görev", "Telefon", "IBAN", "İşe Giriş", "İşten Çıkış", "Aylık Maaş", "Günlük Yevmiye"])
    for p in qs:
        ws.append([p.ad_soyad, p.durum, p.get_personel_turu_display(), p.tc_kimlik_no or "", p.departman or "", p.gorevi or "", p.telefon or "", p.iban or "", p.ise_giris_tarihi, p.isten_cikis_tarihi or "", float(p.aylik_maas or 0), float(p.gunluk_yevmiyesi or 0)])
    _style_workbook(ws)
    return _excel_response(wb, "personel_bilgi_raporu.xlsx")

def _static_image_data_uri(relative_path):
    """Return a data URI for small static images used in printable documents.

    This keeps logos visible when the page is printed or when static files are
    served by gunicorn/Docker instead of Django's development server.
    """
    file_path = settings.BASE_DIR / "static" / relative_path
    try:
        raw = file_path.read_bytes()
    except OSError:
        return ""
    content_type = mimetypes.guess_type(str(file_path))[0] or "image/png"
    encoded = base64.b64encode(raw).decode("ascii")
    return f"data:{content_type};base64,{encoded}"

@login_required
def personel_maas_zarfi(request):
    personeller = tarihte_calisanlar(date.today())
    secili_personel = None
    personel_id = request.GET.get("personel")
    if personel_id:
        secili_personel = get_object_or_404(Personel, id=personel_id)
    return render(request, "personel/maas_zarfi.html", {
        "personeller": personeller,
        "personel": secili_personel,
        "logo_url": request.build_absolute_uri(static("img/logo.png")),
        "logo_data_uri": _static_image_data_uri("img/logo.png"),
    })

@login_required
def istifa_dilekcesi(request):
    personeller = tarihte_calisanlar(date.today())
    secili_personel = None
    dilekce_tarihi = date.today()
    personel_id = request.GET.get("personel")
    tarih = request.GET.get("tarih")
    if personel_id:
        secili_personel = get_object_or_404(Personel, id=personel_id)
    if tarih:
        try: dilekce_tarihi = date.fromisoformat(tarih)
        except ValueError: pass
    return render(request, "personel/istifa_dilekcesi.html", {"personeller": personeller, "personel": secili_personel, "secili_personel": secili_personel, "tarih": dilekce_tarihi})

@login_required
def devamsizlik_tutanagi_sec(request):
    personeller = tarihte_calisanlar(date.today())
    def getp(key):
        pid = request.GET.get(key)
        return get_object_or_404(Personel, id=pid) if pid else None
    return render(request, "personel/devamsizlik_tutanagi_sec.html", {
        "personeller": personeller,
        "secili_personel": getp("personel"),
        "devamsizlik_tarihi": _parse_date(request.GET.get("devamsizlik_tarihi")) or date.today(),
        "isbasi_saati": request.GET.get("isbasi_saati", "08:00"),
        "tutanak_tarihi": _parse_date(request.GET.get("tutanak_tarihi")) or date.today(),
        "tutanak_saati": request.GET.get("tutanak_saati", "09:30"),
        "tanik1": getp("tanik1"),
        "tanik2": getp("tanik2"),
        "isveren_yetkilisi": request.GET.get("isveren_yetkilisi", ""),
        "adres": request.GET.get("adres", ""),
    })

def _parse_date(val):
    if not val: return None
    try: return datetime.strptime(val, "%Y-%m-%d").date()
    except ValueError: return None

@login_required
def devamsizlik_tutanagi(request):
    personel = get_object_or_404(Personel, id=request.GET.get("personel"))
    tanik1 = get_object_or_404(Personel, id=request.GET.get("tanik1"))
    tanik2 = get_object_or_404(Personel, id=request.GET.get("tanik2"))
    sirket = SirketBilgisi.get_solo()
    context = {
        "isveren": {"unvan": sirket.ticari_unvan, "adres": sirket.adres, "sgk_sicil_no": sgk_isyeri_sicil_no_from_departman(personel.departman)},
        "personel": personel,
        "tanik1": tanik1,
        "tanik2": tanik2,
        "devamsizlik_tarihi": _parse_date(request.GET.get("devamsizlik_tarihi")),
        "isbasi_saati": request.GET.get("isbasi_saati"),
        "tutanak_tarihi": _parse_date(request.GET.get("tutanak_tarihi")),
        "tutanak_saati": request.GET.get("tutanak_saati"),
        "isveren_yetkilisi": request.GET.get("isveren_yetkilisi"),
        "adres": request.GET.get("adres", ""),
    }
    return render(request, "personel/devamsizlik_tutanagi.html", context)
