import calendar
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from urllib.parse import quote

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from apps.personel.models import Personel
from apps.puantaj.adapter import PUANTAJ_KODLARI
from apps.puantaj.maas_hakedis_raporlar import maas_hakedis_hesapla
from apps.puantaj.models import BankaOdeme, Puantaj
from apps.puantaj.raporlar import toplu_hakedis_hesapla
from apps.puantaj.services.otomatik_mesai_service import puantaj_kaynakli_mesai_guncelle, puantaj_kaynakli_mesai_temizle
from apps.puantaj.services.puantaj_service import PuantajService
from apps.puantaj.utils import gun_editable_mi

AYLAR = [
    {"value": 1, "label": "Ocak"}, {"value": 2, "label": "Şubat"}, {"value": 3, "label": "Mart"},
    {"value": 4, "label": "Nisan"}, {"value": 5, "label": "Mayıs"}, {"value": 6, "label": "Haziran"},
    {"value": 7, "label": "Temmuz"}, {"value": 8, "label": "Ağustos"}, {"value": 9, "label": "Eylül"},
    {"value": 10, "label": "Ekim"}, {"value": 11, "label": "Kasım"}, {"value": 12, "label": "Aralık"},
]


def _ayda_gorunur_mu(personel, yil, ay):
    gun_sayisi = calendar.monthrange(yil, ay)[1]
    ay_basi = date(yil, ay, 1)
    ay_sonu = date(yil, ay, gun_sayisi)
    if personel.ise_giris_tarihi > ay_sonu:
        return False
    if personel.isten_cikis_tarihi and personel.isten_cikis_tarihi < ay_basi:
        return False
    return True


def _to_float(value):
    if value is None:
        return 0
    return float(Decimal(str(value)))


def _parse_decimal_input(value):
    normalized = str(value or "0").strip().replace(",", ".")
    try:
        result = Decimal(normalized)
    except InvalidOperation as exc:
        raise ValidationError("Banka tutarı sayı olmalıdır. Ondalık için nokta veya virgül kullanabilirsiniz.") from exc
    if not result.is_finite():
        raise ValidationError("Banka tutarı geçerli bir sayı olmalıdır.")
    return result


def _style_workbook(ws):
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 3, 45)
    ws.freeze_panes = "A2"


def _excel_response(wb, filename):
    """Return a real XLSX attachment response.

    Some browsers behave more reliably when the workbook is first saved into
    bytes and the attachment filename is sent with both filename and filename*.
    """
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    safe_filename = filename.replace('"', "")
    response["Content-Disposition"] = (
        f'attachment; filename="{safe_filename}"; filename*=UTF-8\'\'{quote(safe_filename)}'
    )
    response["Content-Length"] = str(output.getbuffer().nbytes)
    response["Cache-Control"] = "no-store"
    return response


@login_required
def puantaj_taslak(request):
    today = date.today()
    ay = int((request.POST if request.method == "POST" else request.GET).get("ay", today.month))
    yil = int((request.POST if request.method == "POST" else request.GET).get("yil", today.year))
    gun_sayisi = calendar.monthrange(yil, ay)[1]
    gunler = list(range(1, gun_sayisi + 1))
    pazar_gunleri = [gun for gun in gunler if date(yil, ay, gun).weekday() == 6]
    personeller = [p for p in Personel.objects.all().order_by("ad_soyad") if _ayda_gorunur_mu(p, yil, ay)]

    if request.method == "POST":
        with transaction.atomic():
            for key, kod in request.POST.items():
                if not key.startswith("gun_"):
                    continue
                try:
                    _, personel_id, gun = key.split("_")
                    personel_id = int(personel_id)
                    gun = int(gun)
                except ValueError:
                    continue
                personel = next((p for p in personeller if p.id == personel_id), None)
                if not personel or not gun_editable_mi(personel, yil, ay, gun):
                    continue
                tarih = date(yil, ay, gun)
                if not kod:
                    Puantaj.objects.filter(personel_id=personel_id, tarih=tarih).delete()
                    puantaj_kaynakli_mesai_temizle(personel, tarih)
                    continue
                if kod not in PUANTAJ_KODLARI:
                    continue
                puantaj, _ = Puantaj.objects.update_or_create(
                    personel_id=personel_id,
                    tarih=tarih,
                    defaults={"kod": kod},
                )
                puantaj_kaynakli_mesai_guncelle(puantaj)
        messages.success(request, "Puantaj kayıtları güncellendi.")
        return redirect(f"/puantaj/?ay={ay}&yil={yil}")

    puantaj_qs = Puantaj.objects.filter(tarih__year=yil, tarih__month=ay, personel__in=personeller)
    kod_map = {(p.personel_id, p.tarih.day): p.kod for p in puantaj_qs}
    ay_basi = date(yil, ay, 1)
    ay_sonu = date(yil, ay, gun_sayisi)
    rows = []
    for personel in personeller:
        day_cells = []
        for gun in gunler:
            day_cells.append({
                "gun": gun,
                "kod": kod_map.get((personel.id, gun), ""),
                "editable": gun_editable_mi(personel, yil, ay, gun),
                "is_pazar": gun in pazar_gunleri,
            })
        ps = PuantajService(personel, ay_basi, ay_sonu)
        rows.append({"personel": personel, "day_cells": day_cells, "toplam": ps.gorunen_toplam_gun()})
    return render(request, "puantaj/puantaj_taslak.html", {
        "ay": ay,
        "yil": yil,
        "aylar": AYLAR,
        "yillar": list(range(today.year - 5, today.year + 6)),
        "gunler": gunler,
        "pazar_gunleri": pazar_gunleri,
        "secili_gun": today.day if yil == today.year and ay == today.month else 1,
        "rows": rows,
        "PUANTAJ_KODLARI": PUANTAJ_KODLARI,
    })


@login_required
def toplu_hakedis_raporu(request):
    baslangic_str = request.GET.get("baslangic")
    bitis_str = request.GET.get("bitis")
    rows = []
    if baslangic_str and bitis_str:
        baslangic = date.fromisoformat(baslangic_str)
        bitis = date.fromisoformat(bitis_str)
        rows = toplu_hakedis_hesapla(baslangic, bitis)
        if request.GET.get("excel") == "1":
            wb = Workbook()
            ws = wb.active
            ws.title = "Toplu Hakediş"
            ws.append(["Personel", "Hakediş Günü", "Resmi Tatil", "Hafta Sonu Çalışması", "Toplam"])
            for r in rows:
                ws.append([
                    r.personel.ad_soyad,
                    _to_float(r.hakedis_gun),
                    _to_float(r.resmi_tatil_gun),
                    _to_float(r.hafta_sonu_gun),
                    _to_float(r.toplam_gun),
                ])
            _style_workbook(ws)
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=5):
                for cell in row:
                    value = Decimal(str(cell.value or 0))
                    cell.number_format = "0" if value == value.to_integral_value() else "0.##"
            return _excel_response(wb, "toplu_hakedis_raporu.xlsx")
    return render(request, "puantaj/toplu_hakedis_raporu.html", {
        "rows": rows,
        "baslangic": baslangic_str,
        "bitis": bitis_str,
    })


@login_required
def maas_hakedis_raporu(request):
    baslangic_str = request.GET.get("baslangic")
    bitis_str = request.GET.get("bitis")
    rows = []
    if baslangic_str and bitis_str:
        rows = maas_hakedis_hesapla(date.fromisoformat(baslangic_str), date.fromisoformat(bitis_str))
        if request.GET.get("excel") == "1":
            wb = Workbook()
            ws = wb.active
            ws.title = "Maaş Hakediş"
            ws.append(["Personel", "Resmi Tatil Gün", "Hakediş Gün", "Hak Edilen Maaş", "Mesai Saat", "Mesai TL", "Prim", "Avans", "İcra", "Net"])
            for r in rows:
                ws.append([
                    r.personel.ad_soyad,
                    _to_float(r.resmi_tatil_gun),
                    _to_float(r.hak_edilen_gun),
                    _to_float(r.hak_edilen_maas),
                    _to_float(r.mesai_saat),
                    _to_float(r.mesai_tutar),
                    _to_float(r.prim_tutar),
                    _to_float(r.avans_tutar),
                    _to_float(r.icra_tutar),
                    _to_float(r.net_odenecek),
                ])
            _style_workbook(ws)
            return _excel_response(wb, "maas_hakedis_raporu.xlsx")
    return render(request, "puantaj/maas_hakedis_raporu.html", {
        "rows": rows,
        "baslangic": baslangic_str,
        "bitis": bitis_str,
    })


@login_required
def banka_elden_listesi(request):
    baslangic_str = request.POST.get("baslangic") or request.GET.get("baslangic")
    bitis_str = request.POST.get("bitis") or request.GET.get("bitis")
    rows = []
    hata = None
    if baslangic_str and bitis_str:
        baslangic = date.fromisoformat(baslangic_str)
        bitis = date.fromisoformat(bitis_str)
        maas_rows = maas_hakedis_hesapla(baslangic, bitis)
        kaydedildi = False
        if request.method == "POST":
            try:
                with transaction.atomic():
                    for r in maas_rows:
                        val = request.POST.get(f"banka_{r.personel.id}", "0") or "0"
                        banka_tutar = _parse_decimal_input(val)
                        if banka_tutar < 0 or banka_tutar > Decimal(str(r.net_odenecek)):
                            raise ValidationError(f"{r.personel.ad_soyad} için banka tutarı net maaşı aşamaz.")
                        BankaOdeme.objects.update_or_create(
                            personel=r.personel,
                            baslangic=baslangic,
                            bitis=bitis,
                            defaults={"banka_tutar": banka_tutar},
                        )
                    kaydedildi = True
            except ValidationError as e:
                hata = str(e)
        if kaydedildi and request.GET.get("excel") != "1":
            messages.success(request, "Banka tutarları kaydedildi.")
            return redirect(f"/puantaj/raporlar/banka-elden/?baslangic={baslangic_str}&bitis={bitis_str}")
        banka_map = {b.personel_id: b for b in BankaOdeme.objects.filter(baslangic=baslangic, bitis=bitis)}
        for r in maas_rows:
            banka = banka_map.get(r.personel.id)
            banka_tutar = banka.banka_tutar if banka else Decimal("0")
            rows.append({
                "personel": r.personel,
                "net_maas": r.net_odenecek,
                "banka": banka_tutar,
                "elden": Decimal(str(r.net_odenecek)) - banka_tutar,
            })
        if request.GET.get("excel") == "1" and not hata:
            wb = Workbook()
            ws = wb.active
            ws.title = "Banka Elden"
            ws.append(["Ad Soyad", "TC Kimlik", "IBAN", "Net Maaş", "Banka Tutarı", "Elden Tutarı"])
            for r in rows:
                ws.append([
                    r["personel"].ad_soyad,
                    r["personel"].tc_kimlik_no or "",
                    r["personel"].iban or "",
                    _to_float(r["net_maas"]),
                    _to_float(r["banka"]),
                    _to_float(r["elden"]),
                ])
            if rows:
                ws.append(["TOPLAM", "", "", f"=SUM(D2:D{len(rows)+1})", f"=SUM(E2:E{len(rows)+1})", f"=SUM(F2:F{len(rows)+1})"])
            _style_workbook(ws)
            for row in ws.iter_rows(min_row=2, min_col=4, max_col=6):
                for cell in row:
                    cell.number_format = '#,##0.00'
            return _excel_response(wb, "banka_elden_raporu.xlsx")
    return render(request, "puantaj/banka_elden_listesi.html", {
        "rows": rows,
        "baslangic": baslangic_str,
        "bitis": bitis_str,
        "hata": hata,
    })


@login_required
def raporlar_index(request):
    return render(request, "puantaj/raporlar_index.html")


# Eski PDF linkleri rapor tasarımı daha sonra yapılacağı için şimdilik HTML rapora yönlenir.
toplu_hakedis_raporu_pdf = toplu_hakedis_raporu
maas_hakedis_raporu_pdf = maas_hakedis_raporu
banka_elden_listesi_pdf = banka_elden_listesi
personel_detay_raporu = maas_hakedis_raporu
personel_detay_raporu_pdf = maas_hakedis_raporu
